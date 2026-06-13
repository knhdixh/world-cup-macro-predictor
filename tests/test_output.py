"""Tests for src/output.py — CSV and Excel output generator."""

from __future__ import annotations

import csv
import shutil
from pathlib import Path
from typing import Any

import openpyxl
import pytest

from src.output import generate_all, generate_csv, generate_excel


@pytest.fixture
def sample_predictions() -> list[dict[str, Any]]:
    """Two prediction dicts as returned by predict_match()."""
    return [
        {
            "team_a": "USA",
            "team_b": "FRA",
            "team_a_xg": 1.8,
            "team_b_xg": 2.1,
            "predicted_score": "2-2",
            "team_a_blended": 0.45,
            "team_b_blended": 0.525,
            "team_a_econ": 78.5,
            "team_b_econ": 82.3,
            "team_a_fifa": 1542.0,
            "team_b_fifa": 1857.0,
        },
        {
            "team_a": "BRA",
            "team_b": "ARG",
            "team_a_xg": 2.5,
            "team_b_xg": 1.9,
            "predicted_score": "3-2",
            "team_a_blended": 0.625,
            "team_b_blended": 0.475,
            "team_a_econ": 65.2,
            "team_b_econ": 58.1,
            "team_a_fifa": 1789.0,
            "team_b_fifa": 1843.0,
        },
    ]


@pytest.fixture
def sample_matches() -> list[dict[str, Any]]:
    """Two schedule match dicts with metadata to merge."""
    return [
        {
            "match_number": 4,
            "group": "D",
            "date": "2026-06-12",
            "team_a": "USA",
            "team_b": "FRA",
            "matchday": 1,
            "status": "upcoming",
        },
        {
            "match_number": 19,
            "group": "J",
            "date": "2026-06-16",
            "team_a": "BRA",
            "team_b": "ARG",
            "matchday": 1,
            "status": "upcoming",
        },
    ]


# ---------------------------------------------------------------------------
# CSV tests
# ---------------------------------------------------------------------------

REQUIRED_COLUMNS = [
    "match_number",
    "group",
    "date",
    "team_a",
    "team_b",
    "predicted_score",
    "team_a_xg",
    "team_b_xg",
    "team_a_econ_score",
    "team_b_econ_score",
    "team_a_fifa_score",
    "team_b_fifa_score",
    "team_a_blended",
    "team_b_blended",
]


def test_csv_headers(
    tmp_path: Path, sample_predictions: list[dict[str, Any]], sample_matches: list[dict[str, Any]]
) -> None:
    """CSV file has all 14 required columns."""
    output_dir = tmp_path / "out"
    csv_path = generate_csv(sample_predictions, sample_matches, output_dir)

    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        assert reader.fieldnames is not None
        assert list(reader.fieldnames) == REQUIRED_COLUMNS


def test_csv_row_count(
    tmp_path: Path, sample_predictions: list[dict[str, Any]], sample_matches: list[dict[str, Any]]
) -> None:
    """Row count matches predictions list length."""
    output_dir = tmp_path / "out"
    csv_path = generate_csv(sample_predictions, sample_matches, output_dir)

    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        next(reader)  # skip header
        rows = list(reader)
    assert len(rows) == len(sample_predictions)


def test_csv_parseable(
    tmp_path: Path, sample_predictions: list[dict[str, Any]], sample_matches: list[dict[str, Any]]
) -> None:
    """File is valid CSV parseable by csv.DictReader."""
    output_dir = tmp_path / "out"
    csv_path = generate_csv(sample_predictions, sample_matches, output_dir)

    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    assert len(rows) == 2
    first = rows[0]
    assert first["match_number"] == "4"
    assert first["group"] == "D"
    assert first["date"] == "2026-06-12"
    assert first["team_a"] == "USA"
    assert first["team_b"] == "FRA"
    assert first["predicted_score"] == "2-2"
    assert float(first["team_a_xg"]) == pytest.approx(1.8)
    assert float(first["team_a_econ_score"]) == pytest.approx(78.5)
    assert float(first["team_a_fifa_score"]) == pytest.approx(1542.0)
    assert float(first["team_a_blended"]) == pytest.approx(0.45)


# ---------------------------------------------------------------------------
# Excel tests
# ---------------------------------------------------------------------------


def test_excel_created(
    tmp_path: Path, sample_predictions: list[dict[str, Any]], sample_matches: list[dict[str, Any]]
) -> None:
    """Excel file created and openable by openpyxl."""
    output_dir = tmp_path / "out"
    excel_path = generate_excel(sample_predictions, sample_matches, output_dir)

    assert Path(excel_path).exists()
    wb = openpyxl.load_workbook(excel_path)
    assert wb is not None
    wb.close()


def test_excel_sheets(
    tmp_path: Path, sample_predictions: list[dict[str, Any]], sample_matches: list[dict[str, Any]]
) -> None:
    """Excel has 2 sheets: 'Predictions' and 'Model Parameters'."""
    output_dir = tmp_path / "out"
    excel_path = generate_excel(sample_predictions, sample_matches, output_dir)

    wb = openpyxl.load_workbook(excel_path)
    assert wb.sheetnames == ["Predictions", "Model Parameters"]
    wb.close()


def test_excel_header_bold(
    tmp_path: Path, sample_predictions: list[dict[str, Any]], sample_matches: list[dict[str, Any]]
) -> None:
    """Header row in 'Predictions' sheet is bold."""
    output_dir = tmp_path / "out"
    excel_path = generate_excel(sample_predictions, sample_matches, output_dir)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Predictions"]
    for cell in ws[1]:
        assert cell.font.bold is True
    wb.close()


# ---------------------------------------------------------------------------
# Directory / integration tests
# ---------------------------------------------------------------------------


def test_auto_create_dir(
    tmp_path: Path, sample_predictions: list[dict[str, Any]], sample_matches: list[dict[str, Any]]
) -> None:
    """Output directory is auto-created if missing."""
    output_dir = tmp_path / "nonexistent" / "nested"
    assert not output_dir.exists()

    csv_path = generate_csv(sample_predictions, sample_matches, output_dir)
    assert Path(csv_path).exists()

    # Clean up so Excel can also test the same scenario fresh
    shutil.rmtree(output_dir)

    excel_path = generate_excel(sample_predictions, sample_matches, output_dir)
    assert Path(excel_path).exists()


# ---------------------------------------------------------------------------
# generate_all integration test
# ---------------------------------------------------------------------------


def test_generate_all_returns_both_paths(
    tmp_path: Path, sample_predictions: list[dict[str, Any]], sample_matches: list[dict[str, Any]]
) -> None:
    """generate_all returns both CSV and Excel paths."""
    output_dir = tmp_path / "all_out"
    csv_path, excel_path = generate_all(sample_predictions, sample_matches, output_dir)

    assert Path(csv_path).exists()
    assert Path(excel_path).exists()
    assert csv_path.endswith("world_cup_predictions.csv")
    assert excel_path.endswith("world_cup_predictions.xlsx")
