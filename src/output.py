"""CSV and Excel output generator for World Cup predictions.

Writes ``world_cup_predictions.csv`` and ``world_cup_predictions.xlsx``
from a list of prediction dicts (output of :func:`predict.predict_match`)
merged with schedule metadata.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font

__all__ = [
    "generate_csv",
    "generate_excel",
    "generate_all",
]

# 14 columns required by the thesis output spec.
CSV_COLUMNS = [
    "match_number",
    "group",
    "date",
    "team_a",
    "team_b",
    "predicted_score",
    "team_a_xg",
    "team_b_xg",
    "team_a_econ_score",
    "team_b_econ_score",
    "team_a_fifa_score",
    "team_b_fifa_score",
    "team_a_blended",
    "team_b_blended",
]


def _merge_prediction_with_match(
    prediction: dict[str, Any], match: dict[str, Any]
) -> dict[str, Any]:
    """Return a flat dict combining prediction values + match metadata.

    Maps prediction keys ``team_a_econ`` → ``team_a_econ_score`` etc.
    """
    return {
        "match_number": match["match_number"],
        "group": match.get("group") if match.get("group") is not None else "",
        "date": match["date"],
        "team_a": prediction["team_a"],
        "team_b": prediction["team_b"],
        "predicted_score": prediction["predicted_score"],
        "team_a_xg": prediction["team_a_xg"],
        "team_b_xg": prediction["team_b_xg"],
        "team_a_econ_score": prediction["team_a_econ"],
        "team_b_econ_score": prediction["team_b_econ"],
        "team_a_fifa_score": prediction["team_a_fifa"],
        "team_b_fifa_score": prediction["team_b_fifa"],
        "team_a_blended": prediction["team_a_blended"],
        "team_b_blended": prediction["team_b_blended"],
    }


def _ensure_dir(output_dir: str | Path) -> Path:
    """Create the output directory if it does not exist."""
    path = Path(output_dir)
    path.mkdir(parents=True, exist_ok=True)
    return path


def generate_csv(
    predictions: list[dict[str, Any]],
    matches: list[dict[str, Any]],
    output_dir: str | Path,
) -> str:
    """Write ``world_cup_predictions.csv`` and return its path.

    Parameters
    ----------
    predictions:
        List of prediction dicts from :func:`predict.predict_match`.
    matches:
        List of schedule dicts with ``match_number``, ``group``, ``date``.
        Must be the same length and order as *predictions*.
    output_dir:
        Directory to write the file into (created if missing).

    Returns
    -------
    str
        Absolute path to the written CSV file.
    """
    out_path = _ensure_dir(output_dir)
    csv_file = out_path / "world_cup_predictions.csv"

    with open(csv_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for pred, match in zip(predictions, matches, strict=True):
            row = _merge_prediction_with_match(pred, match)
            writer.writerow(row)

    return str(csv_file)


def generate_excel(
    predictions: list[dict[str, Any]],
    matches: list[dict[str, Any]],
    output_dir: str | Path,
    *,
    seed_used: int | None = None,
    cutoff_date: str = "2026-06-11",
) -> str:
    """Write ``world_cup_predictions.xlsx`` and return its path.

    Sheet 1 — *Predictions*: all 14 columns, header row bold, auto-width.
    Sheet 2 — *Model Parameters*: blend, noise, seed, cutoff, sources.

    Parameters
    ----------
    predictions:
        List of prediction dicts from :func:`predict.predict_match`.
    matches:
        List of schedule dicts with metadata. Same length/order as
        *predictions*.
    output_dir:
        Directory to write the file into (created if missing).
    seed_used:
        Random seed used during prediction (displayed in Model Parameters).
    cutoff_date:
        Data cutoff date displayed in Model Parameters.

    Returns
    -------
    str
        Absolute path to the written Excel file.
    """
    out_path = _ensure_dir(output_dir)
    excel_file = out_path / "world_cup_predictions.xlsx"

    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------
    # Sheet 1 — Predictions
    # ------------------------------------------------------------------
    ws_pred = wb.active
    ws_pred.title = "Predictions"

    # Header row (bold)
    for col_idx, col_name in enumerate(CSV_COLUMNS, start=1):
        cell = ws_pred.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)

    # Data rows
    for row_idx, (pred, match) in enumerate(zip(predictions, matches, strict=True), start=2):
        row = _merge_prediction_with_match(pred, match)
        for col_idx, col_name in enumerate(CSV_COLUMNS, start=1):
            ws_pred.cell(row=row_idx, column=col_idx, value=row[col_name])

    # Auto-width columns
    for col_idx, col_name in enumerate(CSV_COLUMNS, start=1):
        max_length = len(col_name)
        for row in ws_pred.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
        ws_pred.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_length + 2

    # ------------------------------------------------------------------
    # Sheet 2 — Model Parameters
    # ------------------------------------------------------------------
    ws_param = wb.create_sheet(title="Model Parameters")

    params = [
        ("Blend Weights", "50/50 (FIFA / Economic)"),
        ("Noise Sigma", 0.3),
        ("Seed Used", seed_used if seed_used is not None else "None"),
        ("Cutoff Date", cutoff_date),
        ("Data Sources", "IMF WEO 2026, BIS, FIFA April 2026"),
    ]

    for row_idx, (key, value) in enumerate(params, start=1):
        ws_param.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
        ws_param.cell(row=row_idx, column=2, value=value)

    # Auto-width for parameter sheet
    ws_param.column_dimensions["A"].width = 20
    ws_param.column_dimensions["B"].width = 35

    wb.save(excel_file)
    wb.close()

    return str(excel_file)


def generate_all(
    predictions: list[dict[str, Any]],
    matches: list[dict[str, Any]],
    output_dir: str | Path,
    *,
    seed_used: int | None = None,
    cutoff_date: str = "2026-06-11",
) -> tuple[str, str]:
    """Generate both CSV and Excel outputs.

    Returns
    -------
    tuple[str, str]
        (csv_path, excel_path)
    """
    csv_path = generate_csv(predictions, matches, output_dir)
    excel_path = generate_excel(
        predictions, matches, output_dir, seed_used=seed_used, cutoff_date=cutoff_date
    )
    return csv_path, excel_path
